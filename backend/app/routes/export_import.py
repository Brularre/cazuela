from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db import client
from app.handlers.expenses import normalize as normalize_item
from app.handlers.pantry import normalize as normalize_pantry_item
from app.middleware.auth import require_auth

router = APIRouter(prefix="/dashboard")

_ACCENT_RED = "C0392B"
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color=_ACCENT_RED, end_color=_ACCENT_RED, fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center")


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"


def _autofit(ws) -> None:
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 45)


def _col_map(header_row) -> dict[str, int]:
    return {
        str(h or "").lower().strip().replace("_", " "): i
        for i, h in enumerate(header_row)
    }


def _get(row, col_map: dict, *names):
    for name in names:
        idx = col_map.get(name)
        if idx is not None and idx < len(row):
            val = row[idx]
            if val is not None:
                return val
    return None


def _to_int(val, default=0):
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    try:
        s = str(val).strip()[:10]
        date.fromisoformat(s)
        return s
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@router.get("/export")
def export_data(uid: str = Depends(require_auth)):
    wb = Workbook()
    wb.remove(wb.active)

    # Despensa
    ws = wb.create_sheet("Despensa")
    _write_header(ws, ["item", "categoría", "cantidad_actual", "cantidad_deseada"])
    pantry = (
        client.table("pantry")
        .select("item, category, current_quantity, desired_quantity")
        .eq("user_id", uid)
        .order("item")
        .execute()
    ).data or []
    for row in pantry:
        ws.append([row["item"], row["category"], row["current_quantity"], row["desired_quantity"]])

    # Recetas (flattened — recipe name repeated per ingredient row)
    ws = wb.create_sheet("Recetas")
    _write_header(ws, ["receta", "porciones", "ingrediente", "cantidad", "unidad"])
    recipes = (
        client.table("recipes")
        .select("name, servings, recipe_ingredients(item, quantity, unit)")
        .eq("user_id", uid)
        .order("name")
        .execute()
    ).data or []
    for recipe in recipes:
        ingredients = recipe.get("recipe_ingredients") or []
        if not ingredients:
            ws.append([recipe["name"], recipe["servings"], None, None, None])
        else:
            for ing in ingredients:
                ws.append([
                    recipe["name"],
                    recipe["servings"],
                    ing["item"],
                    ing.get("quantity"),
                    ing.get("unit"),
                ])

    # Lista de compras
    ws = wb.create_sheet("Lista de compras")
    _write_header(ws, ["item", "cantidad", "unidad"])
    lista = (
        client.table("shopping_list")
        .select("item, quantity, unit")
        .eq("user_id", uid)
        .eq("checked", False)
        .order("id")
        .execute()
    ).data or []
    for row in lista:
        ws.append([row["item"], row.get("quantity"), row.get("unit")])

    # Gastos (all time)
    ws = wb.create_sheet("Gastos")
    _write_header(ws, ["fecha", "monto", "moneda", "categoría", "nota"])
    gastos = (
        client.table("expenses")
        .select("date, amount, currency, category, note")
        .eq("user_id", uid)
        .order("date", desc=True)
        .execute()
    ).data or []
    for row in gastos:
        ws.append([
            row["date"],
            float(row["amount"]),
            row.get("currency", "CLP"),
            row["category"],
            row.get("note") or "",
        ])

    # Tareas
    ws = wb.create_sheet("Tareas")
    _write_header(ws, ["tarea", "prioridad", "fecha_límite", "completada"])
    todos = (
        client.table("todos")
        .select("task, priority, due_date, done")
        .eq("user_id", uid)
        .order("created_at")
        .execute()
    ).data or []
    for row in todos:
        ws.append([
            row["task"],
            row.get("priority", "semana"),
            row.get("due_date"),
            "sí" if row["done"] else "no",
        ])

    # Esperando
    ws = wb.create_sheet("Esperando")
    _write_header(ws, ["descripción", "creado_el", "resuelto"])
    waiting = (
        client.table("waiting_on")
        .select("description, created_at, resolved")
        .eq("user_id", uid)
        .order("created_at")
        .execute()
    ).data or []
    for row in waiting:
        ws.append([
            row["description"],
            row["created_at"],
            "sí" if row["resolved"] else "no",
        ])

    for name in wb.sheetnames:
        _autofit(wb[name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cazuela_export.xlsx"},
    )


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    uid: str = Depends(require_auth),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Solo se aceptan archivos .xlsx")

    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="No se pudo leer el archivo")

    results = {}

    # --- Despensa: upsert overwrites all fields including current stock ---
    if "Despensa" in wb.sheetnames:
        rows = list(wb["Despensa"].iter_rows(values_only=True))
        if len(rows) > 1:
            cm = _col_map(rows[0])
            count = 0
            for row in rows[1:]:
                item = str(_get(row, cm, "item") or "").strip()
                if not item:
                    continue
                category = str(_get(row, cm, "categoría", "categoria") or "otros").strip()
                if category not in ("cocina", "baño", "otros"):
                    category = "otros"
                desired = max(_to_int(_get(row, cm, "cantidad deseada"), 1), 0)
                current_raw = _get(row, cm, "cantidad actual")
                current = max(_to_int(current_raw, desired), 0) if current_raw is not None else desired
                client.table("pantry").upsert(
                    {
                        "user_id": uid,
                        "item": normalize_pantry_item(item),
                        "category": category,
                        "desired_quantity": desired,
                        "current_quantity": current,
                    },
                    on_conflict="user_id,item",
                ).execute()
                count += 1
            results["despensa"] = count

    # --- Recetas: update servings + replace ingredients for existing recipes ---
    if "Recetas" in wb.sheetnames:
        rows = list(wb["Recetas"].iter_rows(values_only=True))
        if len(rows) > 1:
            cm = _col_map(rows[0])
            existing = {
                r["name"].lower(): r["id"]
                for r in (
                    client.table("recipes").select("id, name").eq("user_id", uid).execute().data or []
                )
            }
            recipe_ids = {}   # name_lower -> id for recipes touched this import
            wiped = set()     # recipe ids whose old ingredients have been deleted
            count = 0
            for row in rows[1:]:
                name = str(_get(row, cm, "receta") or "").strip()
                if not name:
                    continue
                key = name.lower()
                servings = max(_to_int(_get(row, cm, "porciones"), 2), 1)

                if key not in recipe_ids:
                    if key in existing:
                        recipe_id = existing[key]
                        client.table("recipes").update({"servings": servings}).eq("id", recipe_id).execute()
                    else:
                        result = client.table("recipes").insert(
                            {"user_id": uid, "name": name, "servings": servings}
                        ).execute()
                        if not result.data:
                            continue
                        recipe_id = result.data[0]["id"]
                        existing[key] = recipe_id
                        count += 1
                    recipe_ids[key] = recipe_id

                recipe_id = recipe_ids[key]

                # Delete old ingredients once per recipe on first encounter
                if recipe_id not in wiped:
                    client.table("recipe_ingredients").delete().eq("recipe_id", recipe_id).execute()
                    wiped.add(recipe_id)

                ingredient = str(_get(row, cm, "ingrediente") or "").strip()
                if ingredient:
                    qty = _to_float(_get(row, cm, "cantidad"))
                    unit = str(_get(row, cm, "unidad") or "").strip() or None
                    client.table("recipe_ingredients").insert(
                        {
                            "recipe_id": recipe_id,
                            "item": normalize_item(ingredient),
                            "quantity": qty,
                            "unit": unit,
                        }
                    ).execute()
            results["recetas"] = count

    # --- Lista de compras: update qty/unit if exists, insert if new ---
    if "Lista de compras" in wb.sheetnames:
        rows = list(wb["Lista de compras"].iter_rows(values_only=True))
        if len(rows) > 1:
            cm = _col_map(rows[0])
            existing_items = {
                r["item"]: r["id"]
                for r in (
                    client.table("shopping_list")
                    .select("id, item")
                    .eq("user_id", uid)
                    .eq("checked", False)
                    .execute()
                    .data or []
                )
            }
            count = 0
            for row in rows[1:]:
                item = str(_get(row, cm, "item") or "").strip()
                if not item:
                    continue
                normalized = normalize_pantry_item(item)
                qty = _get(row, cm, "cantidad")
                unit = str(_get(row, cm, "unidad") or "").strip() or None
                if normalized in existing_items:
                    patch = {}
                    if qty is not None:
                        patch["quantity"] = max(_to_int(qty, 0), 0)
                    if unit is not None:
                        patch["unit"] = unit
                    if patch:
                        client.table("shopping_list").update(patch).eq("id", existing_items[normalized]).execute()
                else:
                    insert_data = {"user_id": uid, "item": normalized, "source": "import", "checked": False}
                    if qty is not None:
                        insert_data["quantity"] = max(_to_int(qty, 0), 0)
                    if unit:
                        insert_data["unit"] = unit
                    client.table("shopping_list").insert(insert_data).execute()
                    existing_items[normalized] = ""
                count += 1
            results["lista_compras"] = count

    # --- Tareas: update pending if exists, skip completed, insert new ---
    if "Tareas" in wb.sheetnames:
        rows = list(wb["Tareas"].iter_rows(values_only=True))
        if len(rows) > 1:
            cm = _col_map(rows[0])
            pending = {
                r["task"].lower(): r["id"]
                for r in (
                    client.table("todos").select("id, task").eq("user_id", uid).eq("done", False).execute().data or []
                )
            }
            done = {
                r["task"].lower()
                for r in (
                    client.table("todos").select("task").eq("user_id", uid).eq("done", True).execute().data or []
                )
            }
            count = 0
            for row in rows[1:]:
                task = str(_get(row, cm, "tarea") or "").strip()
                if not task:
                    continue
                completada = str(_get(row, cm, "completada") or "no").strip().lower()
                if completada in ("sí", "si", "yes", "1", "true"):
                    continue
                if task.lower() in done:
                    continue
                priority = str(_get(row, cm, "prioridad") or "semana").strip()
                if priority not in ("hoy", "semana", "mes"):
                    priority = "semana"
                due_date = _to_date(_get(row, cm, "fecha límite"))
                if task.lower() in pending:
                    client.table("todos").update(
                        {"priority": priority, "due_date": due_date}
                    ).eq("id", pending[task.lower()]).execute()
                else:
                    client.table("todos").insert(
                        {"user_id": uid, "task": task, "priority": priority, "due_date": due_date, "done": False}
                    ).execute()
                    pending[task.lower()] = ""
                count += 1
            results["tareas"] = count

    # --- Esperando: insert new unresolved items, skip if already open ---
    if "Esperando" in wb.sheetnames:
        rows = list(wb["Esperando"].iter_rows(values_only=True))
        if len(rows) > 1:
            cm = _col_map(rows[0])
            existing_waiting = {
                r["description"].lower()
                for r in (
                    client.table("waiting_on")
                    .select("description")
                    .eq("user_id", uid)
                    .eq("resolved", False)
                    .execute()
                    .data or []
                )
            }
            count = 0
            for row in rows[1:]:
                desc = str(_get(row, cm, "descripción", "descripcion") or "").strip()
                if not desc:
                    continue
                resuelto = str(_get(row, cm, "resuelto") or "no").strip().lower()
                if resuelto in ("sí", "si", "yes", "1", "true"):
                    continue
                if desc.lower() in existing_waiting:
                    continue
                client.table("waiting_on").insert({"user_id": uid, "description": desc, "resolved": False}).execute()
                existing_waiting.add(desc.lower())
                count += 1
            results["esperando"] = count

    return {"ok": True, "imported": results}
